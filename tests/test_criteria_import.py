import os
import tempfile
import unittest

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None

from app import criteria_import


@unittest.skipIf(Workbook is None, "openpyxl not installed")
class CriteriaImportTests(unittest.TestCase):
    def test_parse_excel(self):
        wb = Workbook()
        ws = wb.active
        # Header
        ws.cell(row=1, column=2, value="Theme")
        ws.cell(row=1, column=3, value="Criteria")
        ws.cell(row=1, column=5, value="Importance (0-100%)")

        ws.cell(row=2, column=2, value="1. Company/Idea Foundations")
        ws.cell(row=3, column=2, value="Science")
        ws.cell(row=4, column=3, value="Feasibility")
        ws.cell(row=4, column=5, value=0.8)

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test.xlsx")
            wb.save(path)
            rows = criteria_import.parse_excel_criteria(path)

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["theme"], "1. Company/Idea Foundations")
        self.assertEqual(row["subtheme"], "Science")
        self.assertEqual(row["name"], "Feasibility")
        self.assertAlmostEqual(row["weight"], 0.8)


if __name__ == "__main__":
    unittest.main()
