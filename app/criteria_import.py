import re
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


HEADER_THEME = "Theme"
HEADER_CRITERIA = "Criteria"


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value, default=0.0):
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        try:
            return float(str(value).strip())
        except (TypeError, ValueError):
            return default


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def build_signal_key(theme: str, subtheme: str, name: str, used: set) -> str:
    base = " ".join([theme or "", subtheme or "", name or ""]).strip()
    slug = slugify(base)[:60]
    key = f"qic_{slug}" if slug else "qic_signal"
    if key not in used:
        used.add(key)
        return key
    suffix = 2
    while f"{key}_{suffix}" in used:
        suffix += 1
    key = f"{key}_{suffix}"
    used.add(key)
    return key


def infer_automation_type(theme: str, subtheme: str, name: str) -> str:
    theme_l = (theme or "").lower()
    name_l = (name or "").lower()
    sub_l = (subtheme or "").lower()

    if "our angle" in theme_l:
        return "manual"

    manual_phrases = [
        "track record",
        "driven to realise success",
        "driven to have a positive impact",
        "sufficient bandwidth",
        "preferential access",
        "understanding of the science",
        "understanding of the market",
        "ability for operational engine",
        "ability to provide follow-ons",
        "governance rights",
        "dilution protection",
    ]

    for phrase in manual_phrases:
        if phrase in name_l:
            return "manual"
    for phrase in manual_phrases:
        if phrase in sub_l:
            return "manual"

    return "ai"


def parse_excel_criteria(path: str) -> List[Dict[str, object]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to parse Excel files")

    wb = load_workbook(path, data_only=True)
    sheet = wb.active

    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        theme = _clean(row[1].value) if len(row) > 1 else ""
        criteria = _clean(row[2].value) if len(row) > 2 else ""
        if theme == HEADER_THEME and criteria == HEADER_CRITERIA:
            header_row = row[1].row
            break

    if header_row is None:
        raise RuntimeError("Could not find header row in Excel sheet")

    current_theme = ""
    current_subtheme = ""
    results: List[Dict[str, object]] = []
    used_keys = set()
    order = 1

    for r in range(header_row + 1, sheet.max_row + 1):
        theme_cell = _clean(sheet.cell(row=r, column=2).value)
        criteria_cell = _clean(sheet.cell(row=r, column=3).value)
        importance = sheet.cell(row=r, column=5).value

        # skip totals
        if theme_cell.lower().startswith("total") or criteria_cell.lower().startswith("total"):
            continue

        if theme_cell and not criteria_cell:
            if re.match(r"^\d+\.", theme_cell):
                current_theme = theme_cell
                current_subtheme = ""
            else:
                current_subtheme = theme_cell
            continue

        if not criteria_cell:
            continue

        note = None
        if theme_cell.startswith("(") and theme_cell.endswith(")"):
            note = theme_cell
        elif theme_cell:
            current_subtheme = theme_cell

        if not current_theme:
            continue

        weight = _to_float(importance, default=1.0)
        name = criteria_cell
        description = None
        if note:
            description = f"{current_subtheme} {note}".strip()
        else:
            description = current_subtheme

        signal_key = build_signal_key(current_theme, current_subtheme, name, used_keys)
        automation_type = infer_automation_type(current_theme, current_subtheme, name)

        results.append(
            {
                "theme": current_theme,
                "subtheme": current_subtheme,
                "name": name,
                "description": description,
                "weight": weight,
                "score_min": 0,
                "score_max": 5,
                "display_order": order,
                "signal_key": signal_key,
                "automation_type": automation_type,
            }
        )
        order += 1

    return results
